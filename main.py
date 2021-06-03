from selenium import webdriver
import time
import base64
import os
import openpyxl
from openpyxl import load_workbook

# 操作するExcelファイルパス
excel_path = "./SearchList.xlsx"

#検索したい単語一覧取得
wb = load_workbook(filename=excel_path, read_only=True)
ws = wb.worksheets[0]
search_column = "A"
search_row = 2

is_get_search_words = True
search_words = []
while is_get_search_words :
    search_cell = search_column + str(search_row)
    search_value = ws[search_cell].value
    if(not search_value is None) :
        search_words.append(search_value)
        search_row += 1
        print(search_value)
    else :
        wb.close()
        is_get_search_words = False

print(search_words)

#search_word = "sieve"

#ドライバーの指定
driver = webdriver.Chrome(executable_path='./chromedriver_win32/chromedriver.exe')

#Google画像検索にアクセス
img_path_list = []

for search_word in search_words:
    driver.get("https://www.google.com/search?q=" + search_word + "&tbm=isch")
    time.sleep(1)
    img = driver.find_element_by_id("islrg").find_element_by_tag_name("img")
    img_data = img.get_attribute("src");

    #画像のデコード
    head, data = img_data.split(',', 1)
    file_ext = "." + head.split(';')[0].split('/')[1]
    plain_data = base64.b64decode(data)

    #画像ファイル名
    foldername = "./download"
    filename = search_word + file_ext
    filepath = os.path.join(foldername,filename)
    img_path_list.append(filepath)
    with open(filepath,"wb") as f:
        f.write(plain_data)

driver.close()
driver.quit()

#検索結果を貼り付け
wb = load_workbook(filename=excel_path)
ws = wb.worksheets[0]
paste_column = "B"
paste_row = 2

for img_path in img_path_list:
    img = openpyxl.drawing.image.Image(img_path)
    paste_cell = paste_column + str(paste_row)
    ws.add_image(img,paste_cell)
    #ws.column_dimensions[paste_column].width = img.width * 0.125
    ws.row_dimensions[paste_row].height = img.height * 0.75
    paste_row += 1

wb.save(excel_path)
wb.close()


